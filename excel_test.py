from openpyxl import load_workbook
from openpyxl import Workbook

list2=[
"YASDAYD_001",	
"YASDAYD_002",
"YASDAYD_003",
"YASDAYD_004",
"YASDAYD_005",
"YASDAYD_006",
"YASDAYD_007",
"YASDAYD_008",
"YASDAYD_009",
"YASDAYD_010",
"YASDAYD_011",
"YASDAYD_012",
"YASDAYD_013",
"YASDAYD_014",
"YASDAYD_015",
"YASDAYD_016",
"YASDAYD_017",
"YASDAYD_018",
"YASDAYD_019",
"YASDAYD_020",
"YASDAYD_021",
"YASDAYD_022",
"YASDAYD_023",
"YASDAYD_024",
"YASDAYD_025",
"YASDAYD_026",
"YASDAYD_027",
"YASDAYD_028",
"YASDAYD_029",
"YASDAYD_030"
]

"""
open函数
参数：
'w'————————————文件打开模式
newline=''—————取消自动隔行
"""
# import csv
# def csv_write(data_old):
#     with open("文件名.csv",'w',newline='') as old_csv:
#         csv.writer(old_csv).writerow(data_old) #写入单行数据
#         old_csv.close()
#     print("文件名.csv数据集--创建成功")

def excel_winte(data_execl):
    excel_path = r"D:\\Github\\python_100\\文件名2.xlsx"  #表格名称
    # wb = Workbook(excel_path)             #创建xlsx表格
    # wb.save(excel_path)                   #保存
    wb = load_workbook(excel_path)        #导入表格，用于下面进行操作
    # wb.create_sheet("sheet2") #创建表名
    print(wb.sheetnames)
    sheets = wb.sheetnames # 获取工作表名称
    ws = wb["Sheet1"] # 获取名为"Sheet1"的工作表对象
    ws = wb.active          #激活
    rows_max = ws.max_row    #获取最大行数
    print(f'当前最大行数{rows_max}')
    ws.append(data_execl)
    wb.save(excel_path)
    print('_________________________分割线_____________________________')
    print(f'当前最大行数{rows_max}')
    wb.close()                            #关闭
    print("%s 保存成功！" % excel_path)





if __name__ == "__main__":
    aa= [1,2,3,4]
    excel_winte(aa)

