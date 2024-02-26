import pymysql
import csv
from openpyxl import load_workbook
from openpyxl import Workbook

excel_path = r"D:\\Github\\python_100\\文件名2.xlsx"  #文件位置


def read_excel_list(number):
    read_sheet_name = 'Sheet'
    wb = load_workbook(excel_path)        #导入表格，用于下面进行操作
    # wb.create_sheet("sheet2") #创建表名
    print(wb.sheetnames)
    sheet_names = wb.sheetnames # 获取工作表名称
    print(sheet_names)
    if read_sheet_name in sheet_names:
        wb_sheet = wb[read_sheet_name] # 切换表名
        print(f'已切换到工作表: {read_sheet_name}')
    else:
        print('表名不存在，结束')
        return
    rows_max = wb_sheet.max_row    #获取最大行数
    print(f'读取{read_sheet_name}中，最大行数{rows_max}')
    LIST_READ_EXCEL = []
    for row in wb_sheet.rows:
        txt = row[number].value
        if txt == None:
            continue
        else:
            LIST_READ_EXCEL.append(txt)
    print(LIST_READ_EXCEL)
    wb.close()     
    return LIST_READ_EXCEL


# 自由调用
if __name__ == "__main__":
    read_excel_list(3)