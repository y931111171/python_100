
import pymysql
import csv
from openpyxl import load_workbook
from openpyxl import Workbook


#参数池
tmp_dict = {
    "YASDAYD_001":"YASDAYD_001_1",	
"YASDAYD_002":"YASDAYD_002_1",
"YASDAYD_003":"YASDAYD_003_1",
"YASDAYD_004":"YASDAYD_004_1",
"YASDAYD_005":"YASDAYD_005_1",
"YASDAYD_006":"YASDAYD_006_1",
"YASDAYD_007":"YASDAYD_007_1",
}

excel_path = r"D:\\Github\\python_100\\文件名2.xlsx"  #文件位置

#读取表格做成列表
def read_excel_list(number):
    read_sheet_name = 'Sheet'
    wb = load_workbook(excel_path)        #导入表格，用于下面进行操作
    # wb.create_sheet("sheet2") #创建表名
    print(wb.sheetnames)
    sheet_names = wb.sheetnames # 获取工作表名称
    print(sheet_names)
    if read_sheet_name in sheet_names:
        wb_sheet = wb[read_sheet_name] # 切换表名
        print(f'已切换到工作表: {read_sheet_name}')
    else:
        print('表名不存在，结束')
        return
    rows_max = wb_sheet.max_row    #获取最大行数
    print(f'读取{read_sheet_name}中，最大行数{rows_max}')
    LIST_READ_EXCEL = []
    for row in wb_sheet.rows:
        txt = row[number].value
        if txt == None:
            continue
        else:
            LIST_READ_EXCEL.append(txt)
    print(LIST_READ_EXCEL)
    wb.close()     
    return LIST_READ_EXCEL

L1 = read_excel_list(3)
print(f'______________读取表格字段,生成列表数据成功___________________________')


#检查文件是否在字典中
def check_in_dict(the_dict,the_key):
    if the_dict.get(the_key):
        print(f'元素在字典中')
        return the_dict.get(the_key)
    else:
        print(f'元素不在字典中')
        return None

"""
open函数
参数：
'w'————————————文件打开模式
newline=''—————取消自动隔行
"""
import csv
def csv_write(data_old):
    #参数化输入
    csv_path=r"D:\\Github\\python_100\\文件名1.csv"  #文件位置
    with open(csv_path,'w',newline='') as old_csv:
        csv.writer(old_csv).writerows(data_old) #写入多行数据
        old_csv.close()
    print("文件名1.csv数据集--创建成功")

def excel_winte(data_execl):
    #参数化输入
    excel_path = r"D:\\Github\\python_100\\文件名2.xlsx"  #文件位置
    target_sheet_name = 'Sheet1'
    # wb = Workbook(excel_path)             #创建xlsx表格
    # wb.save(excel_path)                   #保存
    wb = load_workbook(excel_path)        #导入表格，用于下面进行操作
    # wb.create_sheet("sheet2") #创建表名
    print(wb.sheetnames)
    sheet_names = wb.sheetnames # 获取工作表名称
    print(sheet_names)
    if target_sheet_name in sheet_names:
        wb.active = wb[target_sheet_name] # 切换表名
        print(f'已切换到工作表: {target_sheet_name}')
    else:
        wb.create_sheet(target_sheet_name) #创建表名
        wb.active = wb[target_sheet_name] # 切换表名
        print(f'表名不存在，创建并切换到工作表: {target_sheet_name}')
    ws = wb.active          #激活
    rows_max1 = ws.max_row    #获取最大行数
    print(f'当前最大行数{rows_max1}')
    for i in data_execl:
        ws.append(i)
    wb.save(excel_path)
    print('_________________________分割线_____________________________')
    rows_max2 = ws.max_row 
    print(f'当前表格最大行数{rows_max2}，增加{rows_max2-rows_max1}行数')
    wb.close()                            #关闭
    print("%s 保存成功！" % excel_path)


# 创建库语句
CREATE_SCHEMA_SQL = '''
    create schema wzg charset utf8;
    '''
# 创建表语句
CREATE_TABLE_SQL = '''create table student(
    sno varchar(10) primary key,
    sname varchar(10),
    sage varchar(3),
    ssex enum('男', '女'),
    sadcademy varchar(20),
    sgrade varchar(4),
    sclass varchar(2)
    )default charset=utf8 
    '''
# 插入数据语句
CREATE_student_SQL = '''
insert into student values('2016081111','张三','20','男','软件工程学院','2016','3'),
('2016061111','王杰','21','男','网络工程学院','2016','3'),
('2016071113','周顺','19','男','大气科学学院','2016','3'),
('2017081180','李伟','20','男','软件工程学院','2017','2'),
('2016081201','王丽','20','女','软件工程学院','2016','5')
'''

# # 查询数据
# select_new_table_SQL='''
# select idnew_table,new_tablecol from new_systm.new_table
# where idnew_table=%s 
# '''(idnew_table0)



# 初始化
def init():
    #参数化输入
    tmp_sql_sysm = ['localhost','root','123456','utf8','new_systm']   #参数化连接
    tmp_sql_select1 = ['idnew_table,new_tablecol']
    tmp_sql_parameter= ['jk_0','txt2','idnew_table']

    try:
        DB = pymysql.connect(host=tmp_sql_sysm[0], user=tmp_sql_sysm[1], password=tmp_sql_sysm[2], charset=tmp_sql_sysm[3], database=tmp_sql_sysm[4])
        cursor2 = DB.cursor()
        # cursor2.execute(CREATE_TABLE_SQL)
        # cursor2.execute(CREATE_student_SQL)
        print('数据库进入成功')
        # 查询数据
        tmp_list = []
        for i in L1:
            # print(i)
            idnew_table0 = i
            # print(idnew_table0)
            select_new_table_SQL= f'''select {tmp_sql_select1[0]} from new_systm.new_table where {tmp_sql_parameter[0]}= '{tmp_sql_parameter[1]}' and {tmp_sql_parameter[2]}= '{idnew_table0}' '''   #查询语句加参数
            print(select_new_table_SQL)
            cursor2.execute(select_new_table_SQL)
            result = cursor2.fetchone()
            result2 = check_in_dict(tmp_dict,i)
            print(f'______________打印数据______________\n{result,result2}________________')
            if result == None and result2 == None :
                temp=[ i ,'']
                tmp_list.append(temp)   #如果为数据库和数据池字段空 插入空列表
                print('______________发现数据为空，插入空数据________________________')
                continue
            elif result == None and  result2 != None:
                temp=[ i ,result2]
                tmp_list.append(temp)   #如果数据池不为空 插入空列表
            else:
                tmp_list.append(list(result))  #如果不为空 插入字段
                continue  
        print('_____________________________循环分割线__________________________________')
        print(tmp_list)
        excel_winte(tmp_list)
        csv_write(tmp_list)
            # print(result[0])

    except Exception as e:
        print('——————————————————————执行失败——————————————————————————/n', e)

    finally:
        # cursor1.close()
        cursor2.close()
        DB.close()


# 自由调用
if __name__ == "__main__":
    init()


<<<<<<< HEAD
# text = 'tacfdsy'

# print(text[0])
# print(text[:1])
# print(text[:2])

for i  in range(2,2+1):
    print(i)
=======
>>>>>>> bf4b6b23c55660d3498f8d54fe597c79d58a48cb
